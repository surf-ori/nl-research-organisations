# /// script
# requires-python = ">=3.11"
# dependencies = ["marimo", "requests", "openpyxl", "python-dotenv"]
# ///

import marimo

__generated_with = "0.23.10"
app = marimo.App(width="full")


with app.setup:
    # Setup — imports, constants, and internal helpers
    import marimo as mo
    import json
    from datetime import datetime, timezone
    from pathlib import Path

    import openpyxl
    import requests

    # Cache directory and Zenodo file URL
    DATA_DIR   = Path("data/raw/zenodo")
    ZENODO_URL = "https://zenodo.org/records/18957154/files/nl-orgs-baseline.xlsx?download=1"


@app.function
def load_ror_ids() -> set[str]:
    """Read the cached XLSX baseline and return a set of fully-qualified ROR URL strings.

    Normalises bare ROR IDs (e.g. "02jz4aj89") to "https://ror.org/02jz4aj89".
    Returns an empty set when no XLSX has been downloaded yet.
    """
    xlsx_path = DATA_DIR / "nl-orgs-baseline.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # The ROR column header may be "ROR", "ror", "ror_id", etc.
    ror_col = next((i for i, h in enumerate(headers) if h and str(h).strip().upper() == "ROR"), None)
    if ror_col is None:
        return set()
    ids: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[ror_col]
        if val:
            val = str(val).strip()
            if not val.startswith("https://"):
                val = f"https://ror.org/{val}"
            ids.add(val)
    return ids


@app.function
def fetch(force_refresh: bool = False) -> dict:
    """Download the Zenodo baseline XLSX and cache it to DATA_DIR.

    Skips the download if the file already exists and force_refresh is False.
    Returns {"record_count": int, "fetched_at": str, "source_url": str}.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    meta_path = DATA_DIR / "_metadata.json"
    xlsx_path = DATA_DIR / "nl-orgs-baseline.xlsx"
    if xlsx_path.exists() and not force_refresh and meta_path.exists():
        return json.loads(meta_path.read_text())
    resp = requests.get(ZENODO_URL, timeout=60)
    resp.raise_for_status()
    xlsx_path.write_bytes(resp.content)
    ids = load_ror_ids()
    meta = {
        "fetched_at":   datetime.now(timezone.utc).isoformat(),
        "record_count": len(ids),
        "source_url":   ZENODO_URL,
    }
    meta_path.write_text(json.dumps(meta))
    return meta


@app.cell(hide_code=True)
def header():
    # Header — Zenodo baseline module description and data source overview
    mo.md("""
    ## Zenodo Baseline — NL Research Organisations

    Downloads the curated baseline list of Dutch research organisations from
    [Zenodo](https://zenodo.org/records/18957154) as an XLSX file and extracts
    all ROR IDs to use as the `ori_base_org` flag in the assembled dataset.

    The XLSX is cached at `data/raw/zenodo/nl-orgs-baseline.xlsx`.
    Call `fetch(force_refresh=True)` to re-download.
    """)
    return


@app.cell(hide_code=True)
def summary():
    # Summary — show how many baseline ROR IDs are in the cached XLSX
    xlsx_exists = (DATA_DIR / "nl-orgs-baseline.xlsx").exists()
    content = (
        mo.md(f"**{len(load_ror_ids())} ROR IDs** in the Zenodo baseline")
        if xlsx_exists
        else mo.callout(
            mo.md("No XLSX cached yet — run `fetch()` or use the Pipeline Stages tab."),
            kind="warn",
        )
    )
    content
    return


if __name__ == "__main__":
    app.run()
